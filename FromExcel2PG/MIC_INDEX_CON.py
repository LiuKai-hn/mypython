import csv
import os
import openpyxl
import time
from GSL.FromExcel2PG.shi_jie_jing_ji.ShiJieJingJiDataInsert import getConn


def Xlsx2CSV(infile_path, outfile_path):
    # 打开文件，xlrd.open_workbook()，函数中参数为文件路径，分为相对路径和绝对路径
    workbook = openpyxl.load_workbook(infile_path)
    # 获取sheet内容
    # 按索引号获取sheet内容
    sheet1_content1 = workbook['Sheet1']  # sheet索引从0开始

    # # 获取所有工作表名称
    # sheet_names = workbook.sheetnames
    #
    # # 打印所有工作表名称
    # print(sheet_names)
    lists = ["F_LEVEL_ONE_CODE," +
             "F_LEVEL_ONE_NAME," +
             "F_LEVEL_TWO_CODE," +
             "F_LEVEL_TWO_NAME," +
             "F_LEVEL_THR_CODE," +
             "F_LEVEL_THR_NAME," +
             "F_CODE," +
             "F_NAME," +
             "F_CODE_FROM_LD_CODE," +
             "F_STATUS"]
    # 遍历 sheet 页的每一行
    for row in sheet1_content1.iter_rows():
        # 遍历每一行中的每一个单元格
        line = []
        # 遍历CSV文件中的每一行,如果没有四级指标，需要在三级指标的基础上+01,如果有则用原来的
        # 如果四级标签不存在 四级标签的中文名称copy三级标签
        if row[6].value is None or len(row[6].value) == 0:
            row[6].value = row[4].value + "01"
        if row[7].value is None or len(row[7].value) == 0:

        for cell in row:
            line.append(str(cell.value))
        line.append("1")  # status
        lists.append(",".join(line))
    # 判断目录是否存在，若不存在，则创建
    # dir_path = os.path.dirname(outfile_path)
    # if not os.path.exists(dir_path):
    #     os.makedirs(dir_path)
    # 如果文件已经存在
    if os.path.exists(outfile_path):
        os.remove(outfile_path)
    # 将数据写入文件
    with open(outfile_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lists))


def InsertData(abs_file_path):
    # 创建游标对象
    conn = getConn()
    cur = conn.cursor()
    # 删除F_CODE原有的数据，避免多次插入
    truncate_sql = "truncate" + " table MIC_INDEX_CON"
    print("execute sql ==>" + truncate_sql)

    # 执行查询语句
    cur.execute(truncate_sql)
    # 提交事务
    conn.commit()
    time.sleep(1)

    # 执行copy from语句
    print("UpsertData copy from ==> " + abs_file_path)
    with open(abs_file_path, 'r', encoding='utf-8') as f:
        next(f)  # 跳过标题行
        columns = ('f_level_one_code', 'f_level_one_name', 'f_level_two_code', 'f_level_two_name', 'f_level_thr_code',
                   'f_level_thr_name', 'f_code', 'f_name', 'f_code_from_ld_code', 'f_status')
        cur.copy_from(f, 'mic_index_con', sep=',', columns=columns)  # 将数据从文件中复制到数据库表中
    # 提交事务
    conn.commit()
    time.sleep(1)
    cur.close()
    conn.close()


if __name__ == '__main__':
    Xlsx2CSV("index.xlsx", "index.csv")
    InsertData("index.csv")
