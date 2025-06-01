#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# --------------------------------
# Author: liu kai
# CreateTime: 2023/5/23 14:36
# Description: 主要解决 国民经济、行业指标 的数据入库，需要注意的是，插入的字段要比世界经济少一个字段 f_area
# Question:


import os
import openpyxl
import time

from GSL.FromExcel2PG.shi_jie_jing_ji.ShiJieJingJiDataInsert import getConn


def Xlsx2CSV(infile_path, outfile_path):
    # 打开文件，xlrd.open_workbook()，函数中参数为文件路径，分为相对路径和绝对路径
    workbook = openpyxl.load_workbook(infile_path)
    # 获取sheet内容
    # 按索引号获取sheet内容
    sheet1_content1 = workbook['Sheet1']  # sheet索引从0开始

    # # 获取所有工作表名称
    # sheet_names = workbook.sheetnames
    #
    # # 打印所有工作表名称
    # print(sheet_names)

    # 获取 龙盾和工商联映射关系
    my_dict = getLongDunData2GSLF_CODEMap()

    lists = []
    # 定义一个set 用于存放 f_code ,避免多次copy from 时出现重复
    gsl_f_code_set = set()
    # 遍历 sheet 页的每一行
    for row in sheet1_content1.iter_rows(min_row=2):
        # 遍历每一行中的每一个单元格
        gsl_code_name = my_dict.get(row[0].value)
        if gsl_code_name is None or len(gsl_code_name) == 0:
            print(row[0].value + " 没有匹配到龙盾对应的gsl F_CODE")
            # continue
        else:
            f_time = row[2].value

            if '月' in f_time:
                f_type_code = 'A'
                f_type = '月度数据'
            elif '季' in f_time:
                f_type_code = 'B'
                f_type = '季度数据'
            else:
                f_type_code = 'C'
                f_type = '年度数据'

            line = str(f_type_code) + "," + str(f_type) + "," + str(gsl_code_name[0]) + "," + str(gsl_code_name[1]) + "," + str(row[2].value) + "," + str(row[3].value) + ",1"
            # print("line===>" + line)
            lists.append(line)
    # f_code 去重
    for line in lists:
        gsl_f_code_set.add(line.split(",")[2])

    # print(gsl_f_code_set)
    # 判断目录是否存在，若不存在，则创建
    dir_path = os.path.dirname(outfile_path)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    # 如果文件已经存在
    if os.path.exists(outfile_path):
        os.remove(outfile_path)
    # 将数据写入文件
    with open(outfile_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lists))
    return str(gsl_f_code_set).replace("{", "(").replace("}", ")")


def getLongDunData2GSLF_CODEMap():
    conn = getConn()
    cur = conn.cursor()
    # 查询 F_CODE 原有的数据
    select_sql = "select F_CODE_FROM_LD_CODE,F_CODE,F_NAME " \
                 "from mic_index_con " \
                 "where F_CODE_FROM_LD_CODE is not null "

    cur.execute(select_sql)
    # 获取查询结果
    rows = cur.fetchall()
    # 如果输入的 指标中文名称有误，则退出
    if len(rows) == 0:
        return
    my_dict = {}
    for row in rows:
        my_dict[row[0]] = (row[1], row[2])

    # 遍历字典的键值对
    # for key, value in my_dict.items():
    #     print(key, value)

    return my_dict


def UpsertData(f_codeset, abs_file_path):
    # 创建游标对象
    conn = getConn()
    cur = conn.cursor()
    # 删除F_CODE原有的数据，避免多次插入
    delete_sql = "DELETE from " + "macro_eco_index where f_code in $f_codes$".replace('$f_codes$', f_codeset)
    print("execute sql ==>" + delete_sql)

    # 执行查询语句
    cur.execute(delete_sql)
    # 提交事务
    conn.commit()
    time.sleep(1)

    # 执行copy from语句
    print("UpsertData copy from ==> " + abs_file_path)
    with open(abs_file_path, 'r', encoding='utf-8') as f:
        # next(f)  # 跳过标题行 已经过滤掉了
        columns = ('f_typecode', 'f_type', 'f_code', 'f_name', 'f_time', 'f_data', 'status')
        cur.copy_from(f, 'macro_eco_index', sep=',', columns=columns)  # 将数据从文件中复制到数据库表中
    # 提交事务
    conn.commit()
    time.sleep(1)
    cur.close()
    conn.close()


if __name__ == '__main__':
    f_code_set = Xlsx2CSV(r"E:\work\gsl\国民经济指标\国民经济指标.xlsx", r"E:\work\gsl\output\guo_min_jing_ji\guo_min_jing_ji.csv")
    UpsertData(f_code_set, r"E:\work\gsl\output\guo_min_jing_ji\guo_min_jing_ji.csv")

