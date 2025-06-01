import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 数据库连接配置
db_config = {
    'host': 'localhost',  # 替换为你的MySQL主机地址
    'user': 'root',  # 替换为你的MySQL用户名
    'password': '123456',  # 替换为你的MySQL密码
    'database': 'baobiao',  # 替换为你的数据库名
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# 需要查询的报表类型
REPORT_TYPES = ['迁入迁出', '服务包','aaa']
reports_str = ", ".join([f"'{r}'" for r in REPORT_TYPES])


def get_recent_dates(days):
    """获取数据库中指定报表类型的最近几个日期（降序排列）"""
    try:
        connection = pymysql.connect(**db_config)
        with connection.cursor() as cursor:
            sql = """
                SELECT DISTINCT compute_dt 
                FROM area_monitor_baobiao_statstic 
                WHERE baobiao_name IN (reports_str)
                ORDER BY compute_dt DESC 
                LIMIT days
            """.replace('days', days).replace('reports_str', reports_str)

            # print("get_data_sql", sql)

            cursor.execute(sql)
            return [row['compute_dt'] for row in cursor.fetchall()]
    finally:
        if connection:
            connection.close()


def fetch_report_data(dates):
    """获取报表数据"""
    try:
        connection = pymysql.connect(**db_config)
        with connection.cursor() as cursor:
            # 构建IN条件字符串
            dates_str = ", ".join([f"'{d}'" for d in dates])

            sql = f"""
                SELECT baobiao_name, compute_dt, compute_cnt 
                FROM area_monitor_baobiao_statstic 
                WHERE baobiao_name IN ({reports_str})
                AND compute_dt IN ({dates_str})
                ORDER BY baobiao_name, compute_dt DESC
            """
            cursor.execute(sql)

            # 整理数据结构：{报表名: {日期: 值}}
            report_data = {}
            for row in cursor.fetchall():
                name = row['baobiao_name']
                if name not in report_data:
                    report_data[name] = {}
                report_data[name][row['compute_dt']] = row['compute_cnt']

            print(report_data)
            return report_data
    finally:
        if connection:
            connection.close()


def export_to_excel(report_data, dates, filename="报表数据.xlsx"):
    """导出数据到Excel并设置格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "报表汇总"

    # 定义样式
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橘黄色
    red_font = Font(color="FF0000")  # 红色
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    # 定义边框样式（实线）
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 生成表头
    headers = ["表名", dates[0], "变化量"] + dates[1:]
    ws.append(headers)

    # 设置表头样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # 添加数据
    for report in REPORT_TYPES:
        row = [report]
        data = report_data.get(report, {})

        # 添加最新日期的值
        latest_value = data.get(dates[0], "无数据")
        row.append(latest_value)

        # 计算变化量（最新日期 - 次新日期）
        if len(dates) > 1 and dates[0] in data and dates[1] in data:
            change = int(data[dates[0]]) - int(data[dates[1]])
            row.append(change)
        else:
            row.append("无数据")

        # 添加剩余日期的值
        for date in dates[1:]:
            row.append(data.get(date, "无数据"))

        ws.append(row)

        center_alignment = Alignment(horizontal="center")
        # 设置当前行的边框和样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

            # 如果是变化量列且为负数，设置为红色
            if col == 3 and isinstance(cell.value, int) and cell.value < 0:
                cell.font = red_font

    # 设置列宽自适应
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 保存文件
    wb.save(filename)
    print(f"Excel文件已生成: {filename}")

if __name__ == '__main__':
    # 获取最近5天的日期
    recent_dates = get_recent_dates('5')
    print(f"使用的日期: {recent_dates}")

    # 获取报表数据
    report_data = fetch_report_data(recent_dates)

    # 导出到Excel
    export_to_excel(report_data, recent_dates)